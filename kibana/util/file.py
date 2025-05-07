import pytesseract
from PIL import ImageGrab
import pyautogui
import pygetwindow as gw
import time
import psutil
import pyperclip
import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string
import os

import win32com.client
from openpyxl.utils import get_column_letter

def open_excel_find_invoice_id(file_path, column_title):

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True 

    try:
        workbook = excel.Workbooks.Open(file_path)
        time.sleep(1.5) 
        focus_excel_full_screen(file_path)
        if excel.ProtectedViewWindows.Count > 0:
            for pv in excel.ProtectedViewWindows:
                if file_path in pv.Workbook.FullName:
                    print(f"File '{file_path}' is in Protected View. Enabling editing...")
                    pv.Edit()
                    time.sleep(2)


        for sheet in workbook.Sheets:
            used_range = sheet.UsedRange  
            for row in range(1, used_range.Rows.Count + 1):
                for col in range(1, used_range.Columns.Count + 1):
                    cell_value = str(used_range.Cells(row, col).Value).strip() if used_range.Cells(row, col).Value else ""
                    if cell_value.lower() == column_title:
                        column_letter = get_column_letter(col) 
                        print(f"Found {column_title} at Sheet: {sheet.Name}, Column: {column_letter}")
                        return column_letter 

        print("No cell with 'Invoice_ID' found in the file.")
        return None

    except Exception as e:
        print(f"Error: {e}")
        return None

    finally:
        pass  

def find_first_empty_cell_in_row(file_path, row_num):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    row_values = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

    last_filled_col = 0
    for col_idx, value in enumerate(row_values, start=1):
        if value not in (None, ""): 
            last_filled_col = col_idx

    if last_filled_col > 0:
        first_empty_col_letter = openpyxl.utils.get_column_letter(last_filled_col + 1)
        second_empty_col_letter = openpyxl.utils.get_column_letter(last_filled_col + 2)
        return first_empty_col_letter, second_empty_col_letter

    return openpyxl.utils.get_column_letter(1), openpyxl.utils.get_column_letter(2)

def is_excel_file_open(file_name):
    for process in psutil.process_iter(['name', 'pid']):
        if process.info['name'] and 'EXCEL' in process.info['name'].upper():
            excel_windows = gw.getWindowsWithTitle(file_name)
            for window in excel_windows:
                if file_name in window.title:
                    print(f"Excel file '{file_name}' is already open.")
                    window.activate()
                    return True
    return False

    chrome_running = False

    for process in psutil.process_iter(['name']):
        if process.info['name'] and 'chrome' in process.info['name'].lower():
            chrome_running = True
            break

    if not chrome_running:
        print("Chrome is not running.")
        return False

    chrome_windows = gw.getWindowsWithTitle(tab_name)
    
    for window in chrome_windows:
        if tab_name in window.title:
            print(f"Tab '{tab_name}' found in Chrome.")

            window.activate()
            time.sleep(1)

            screenshot = pyautogui.screenshot()
            extracted_text = pytesseract.image_to_string(screenshot)

            if url in extracted_text:
                print(f"URL '{url}' found in the Chrome tab.")
                return True
            else:
                print(f"Tab '{tab_name}' found but URL does not match.")

    print(f"Chrome is open, but no tab with '{tab_name}' and URL '{url}' found.")
    return False

def open_excel(fileName):
    excel_file_path = fileName
    
    if not is_excel_file_open(fileName):
        print(f"Opening Excel file: {fileName}")
        os.startfile(excel_file_path)
        time.sleep(2)
    
    try:
        button_location = pyautogui.locateOnScreen("image/enable_edit.png", confidence=0.8)
        if button_location:
            print("Enable Editing button detected! Clicking it...")
            x, y = pyautogui.center(button_location)
            pyautogui.click(x, y)
        else:
            print("No Enable Editing button found. Excel is ready to use.")
    except pyautogui.ImageNotFoundException:
        print("No 'Enable Editing' button detected. Excel is already in editing mode.")

def insert_value(file_name,cell, start_index, value):

    cell_address = f"{cell}{start_index}"
    pyperclip.copy(cell_address)
    pyautogui.hotkey("ctrl", "g")

    time.sleep(1)

    print(f'cell: {cell_address}')
    pyautogui.hotkey("ctrl","v")
    time.sleep(1)

    pyautogui.press("enter")
    time.sleep(1)
    pyperclip.copy(value)
    pyautogui.hotkey("ctrl","v")

    time.sleep(1)  

def count_excel_rows(file_path, start_row):
    df = pd.read_excel(file_path, header=None)

    df = df.dropna(how='all')

    row_count = len(df.iloc[start_row:])
    return row_count

def check_if_excel_running():
    windows = gw.getAllTitles()
    for window in windows:
        if "Excel" in window:
            return window

    return None

def focus_excel_full_screen(file_name):
    print("Switch to excel!")
    excel_window = check_if_excel_running()
    
    if excel_window:
        excel_window = gw.getWindowsWithTitle(excel_window)[0]
        excel_window.activate()
        excel_window.restore()
        excel_window.maximize() 
        time.sleep(1)

def save_file(file_name):
    pyautogui.hotkey("ctrl", "s")

def close_file():
    pyautogui.hotkey("alt", "f4")

def fill_no_call_in_column_s_string(file_path, cell_name):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        col_index = column_index_from_string(cell_name)
        max_row = sheet.max_row

        for row_num in range(3, max_row - 1):
            cell = sheet.cell(row=row_num, column=col_index)
            if cell.value is None or cell.value == "":
                cell.value = "No callback"

        workbook.save(file_path)
        print(f"Empty cells in column {cell_name} (rows 3 to {max_row}) filled with 'No callback' in '{file_path}'.")

    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
    except Exception as e:
        print(f"An error occurred: {e}")