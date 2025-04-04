import pytesseract
from PIL import ImageGrab
import pyautogui
import pygetwindow as gw
import time
import psutil
import pyperclip
import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string
import os


def is_excel_file_open(file_name):
    for process in psutil.process_iter(['name', 'pid']):
        if process.info['name'] and 'EXCEL' in process.info['name'].upper():
            excel_windows = gw.getWindowsWithTitle(file_name)
            for window in excel_windows:
                if file_name in window.title:
                    print(f"Excel file '{file_name}' is already open.")
                    window.activate()
                    return True
    return False

    chrome_running = False

    for process in psutil.process_iter(['name']):
        if process.info['name'] and 'chrome' in process.info['name'].lower():
            chrome_running = True
            break

    if not chrome_running:
        print("Chrome is not running.")
        return False

    chrome_windows = gw.getWindowsWithTitle(tab_name)
    
    for window in chrome_windows:
        if tab_name in window.title:
            print(f"Tab '{tab_name}' found in Chrome.")

            window.activate()
            time.sleep(1)

            screenshot = pyautogui.screenshot()
            extracted_text = pytesseract.image_to_string(screenshot)

            if url in extracted_text:
                print(f"URL '{url}' found in the Chrome tab.")
                return True
            else:
                print(f"Tab '{tab_name}' found but URL does not match.")

    print(f"Chrome is open, but no tab with '{tab_name}' and URL '{url}' found.")
    return False

def open_excel(fileName):
    """Opens an Excel file if not already open, or brings it to focus."""

    excel_file_path = fileName
    
    if not is_excel_file_open(fileName):
        print(f"Opening Excel file: {fileName}")
        os.startfile(excel_file_path)
        # focus_excel_full_screen(excel_file_path)
        time.sleep(2)
    
    try:
        button_location = pyautogui.locateOnScreen("image/enable_edit.png", confidence=0.8)
        if button_location:
            print("Enable Editing button detected! Clicking it...")
            x, y = pyautogui.center(button_location)
            pyautogui.click(x, y)
        else:
            print("No Enable Editing button found. Excel is ready to use.")
    except pyautogui.ImageNotFoundException:
        print("No 'Enable Editing' button detected. Excel is already in editing mode.")

def insert_value(file_name,cell, start_index, value):

    pyautogui.hotkey("ctrl", "g")
    time.sleep(1)

    cell_address = f"{cell}{start_index}"
    pyautogui.write(cell_address, interval=0.1)
    time.sleep(1)

    pyautogui.press("enter")
    time.sleep(1)

    pyautogui.write(value, interval=0.1)
    time.sleep(1)

    pyautogui.press("enter")

    time.sleep(1)  

def count_excel_rows(file_path, start_row):
    df = pd.read_excel(file_path, header=None)

    df = df.dropna(how='all')

    row_count = len(df.iloc[start_row:])
    print("total row: ", row_count)
    return row_count

def check_if_excel_running():
    """Check if an Excel window is open."""
    windows = gw.getAllTitles()
    for window in windows:
        if "Excel" in window:
            return window

    return None

def focus_excel_full_screen(file_name):
    excel_window = check_if_excel_running()
    
    if excel_window:
        print(f"Excel is already open: {excel_window}")
        excel_window = gw.getWindowsWithTitle(excel_window)[0]
        excel_window.activate()
        excel_window.restore()
        excel_window.maximize() 
        time.sleep(1)

def save_file(file_name):
    pyautogui.hotkey("ctrl", "s")

def close_file():
    pyautogui.hotkey("alt", "f4")


def fill_no_call_in_column_s_string(file_path, cell_name):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        col_index = column_index_from_string(cell_name)
        max_row = sheet.max_row

        for row_num in range(3, max_row - 1):
            cell = sheet.cell(row=row_num, column=col_index)
            if cell.value is None or cell.value == "":
                cell.value = "No callback"

        workbook.save(file_path)
        print(f"Empty cells in column {cell_name} (rows 3 to {max_row}) filled with 'No callback' in '{file_path}'.")

    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
    except Exception as e:
        print(f"An error occurred: {e}")